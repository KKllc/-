# -*- coding: utf-8 -*-
"""
Created on Tue Jul  9 10:05:06 2019

@author: hp
"""

import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook

wb = load_workbook("D:/Tencent/QQ/远场r=30cm/无结构-r-30-远场(数据处理版).xlsx")
sheets = wb.sheetnames
print(sheets)

sheet1 = wb[sheets[5]]#无结构振幅表格
sheet2 = wb[sheets[1]]#无结构相位表格

wb1 = load_workbook("D:/Tencent/QQ/远场r=30cm/有结构-空心-r-30-远场（数据处理版）.xlsx")
sheets_1 = wb1.sheetnames
print(sheets_1)

sheet3 = wb1[sheets_1[5]]#有结构振幅表格
sheet4 = wb1[sheets_1[1]]#有结构相位表格

num_list1=np.zeros((120,51))#无结构振幅数组
num_list2=np.zeros((120,51))#无结构相位数组
num_list3=np.zeros((120,51))#有结构振幅数组
num_list4=np.zeros((120,51))#有结构相位数组
num_list5=np.zeros((120,51))#无结构振幅平方数组
num_list6=np.zeros((120,51))#无结构实部数组
num_list7=np.zeros((120,51))#无结构虚部数组
num_list8=np.zeros((120,51))#有结构振幅平方数组
num_list9=np.zeros((120,51))#有结构实部数组
num_list10=np.zeros((120,51))#有结构虚部数组
num_list11=np.zeros((120,51))#散射场振幅平方数组


i=0
j=0
for i in range(0,120):
    for j in range(0,51):
        num_list1[i][j] = sheet1.cell(row=i+1,column=j+1).value
        j+=1
    i+=1
#print(num_list1)

i=0
j=0
for i in range(0,120):
    for j in range(0,51):
        num_list5[i][j] = num_list1[i][j]*num_list1[i][j]
        j+=1
    i+=1

i=0
j=0
for i in range(0,120):
    for j in range(0,51):
        num_list2[i][j] = sheet2.cell(row=i+1,column=j+1).value
        j+=1
    i+=1
#print(num_list1)


m=0
n=0
for m in range(0,120):
    for n in range(0,51):
        num_list6[m][n]=num_list1[m][n]*math.cos(num_list2[m][n]*math.pi/180)
        n+=1
    m+=1
#print(num_list1)

m=0
n=0
for m in range(0,120):
    for n in range(0,51):
        num_list7[m][n]=num_list1[m][n]*math.sin(num_list2[m][n]*math.pi/180)
        n+=1
    m+=1 

i=0
j=0
for i in range(0,120):
    for j in range(0,51):
        num_list3[i][j] = sheet3.cell(row=i+1,column=j+1).value
        j+=1
    i+=1
#print(num_list3)

i=0
j=0
for i in range(0,120):
    for j in range(0,51):
        num_list4[i][j] = sheet4.cell(row=i+1,column=j+1).value
        j+=1
    i+=1
#print(num_list4)

m=0
n=0
for m in range(0,120):
    for n in range(0,51):
        num_list9[m][n]=num_list3[m][n]*math.cos(num_list4[m][n]*math.pi/180)
        n+=1
    m+=1
#print(num_list1) 

m=0
n=0
for m in range(0,120):
    for n in range(0,51):
        num_list10[m][n]=num_list3[m][n]*math.sin(num_list4[m][n]*math.pi/180)
        n+=1
    m+=1

m=0
n=0
for m in range(0,120):
    for n in range(0,51):
        num_list11[m][n]=(num_list9[m][n]-num_list6[m][n])*(num_list9[m][n]-num_list6[m][n])+(num_list10[m][n]-num_list7[m][n])*(num_list10[m][n]-num_list7[m][n])
        n+=1
    m+=1



data1 = pd.DataFrame(num_list5)
data2 = pd.DataFrame(num_list11)  

with pd.ExcelWriter("D:/Tencent/QQ/远场r=30cm/write_r_20.xlsx") as writer:
    data1.to_excel(writer,sheet_name='pi,pi')
    data2.to_excel(writer,sheet_name='ps,ps')

writer.save()
writer.close