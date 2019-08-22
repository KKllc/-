# -*- coding: utf-8 -*-
"""
Created on Mon Jul  8 11:04:13 2019

@author: 888
"""

import pandas as pd
import numpy as np
import math
from openpyxl import load_workbook

'''
data = pd.read_excel("F:/data/2019.07.02/conduct/无结构-r-20-远场.xlsx",sheet_name='Sheet2');

df = pd.read_excel("F:/data/2019_07_08/pandas_data_conduction/hello.xlsx");
data = df.ix[[1,2],['1716','1732']]
print (data)
'''
wb = load_workbook("D:/Computation Physics/hello.xlsx")
sheets = wb.sheetnames
print(sheets)

sheet1 = wb[sheets[0]]
sheet2 = wb[sheets[1]]


num_list1=np.zeros((4,5))
'''
i = 0
j = 0
for row in sheet1.rows:
    j=0
    for cell in row:
        num_list1[i][j] = cell.value
        j+=1
    i+=1
'''
print(num_list1)

print(sheet1.cell(row=3,column=4).value)

i = 0
j = 0
for i in range(0,4):
    for j in range(0,5):
        num_list1[i][j] = sheet1.cell(row=i+1,column=j+1).value
        j+=1
    i+=1
print(num_list1)  

m = 0
n = 0

for m in range(0,4):
    for n in range(0,5):
        num_list1[m][n]=math.cos(num_list1[m][n]*math.pi)
        n+=1
    m+=1
print(num_list1)




data = pd.DataFrame(num_list1)  

writer = pd.ExcelWriter("D:/Computation Physics/write_in.xlsx")
data.to_excel(writer,sheet_name='Sheet1')  
writer.save()
writer.close




